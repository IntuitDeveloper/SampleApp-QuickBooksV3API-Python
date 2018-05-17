import os
import openpyxl

APP_ROOT = os.path.dirname(os.path.abspath(__file__))   # refers to application_top
APP_STATIC = os.path.join(APP_ROOT, '../Leads.xlsx')

excel_file = APP_STATIC

def load_excel():
    """Excel is loaded into a list of dictionaries"""
    wb = openpyxl.load_workbook(filename=excel_file)
    ws = wb['data']

    # Parse data from Excel file
    dict_list = []
    keys = {'A': 'Id', 'B': 'Full Name', 'C': 'Email', 'D': 'Phone'}
    for row in ws.iter_rows(min_row=2, max_col=4):
        d = {}
        for cell in row:
            if cell.column == 'A' and cell.value == None:
                break
            elif cell.column != 'A' and cell.value == None:
                d[keys[cell.column]] = ''
            else:
                d[keys[cell.column]] = str(cell.value)

        if len(d.keys()) != 0:
            dict_list.append(d)
    return dict_list

def remove_lead(customer_list, customer_id):
    """Update lead in excel worksheet and working dictionary; Removing from excel leaves a blank row in the excel"""
    wb = openpyxl.load_workbook(filename=excel_file)
    ws = wb['data']
    for row in ws.iter_rows(min_row=2, max_col=4):
        for cell in row:
            if cell.value == None or cell.column == 'B' or cell.column == 'C' or cell.column == 'D':
                continue
            elif cell.column == 'A' and str(cell.value) == customer_id:
                cell_row = str(cell.row)
                cell.value = None
                ws['B' + cell_row].value = None
                ws['C' + cell_row].value = None
                ws['D' + cell_row].value = None
                break
    wb.save(filename=excel_file)

    # Remove selected lead from excel sheet
    for customer in customer_list:
        if customer['Id'] == customer_id:
            customer_list.pop(customer_list.index(customer))
    return customer_list
